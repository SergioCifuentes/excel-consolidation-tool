from openpyxl import load_workbook, utils

MASTER_PATH = r"master.xlsx"        #Master Excel File for Consolidation


class ExcelManager(object):

    def __new__(cls):
        #if instance already exists it doesn't create a new one  
        if not hasattr(cls, 'instance'):
            cls.instance = super(ExcelManager, cls).__new__(cls)
        return cls.instance

    def __getWorkbook(self, file):
        #get workbook from file with openpyxl
        try:
            workbook = load_workbook(file)
            return workbook
        except utils.exceptions.InvalidFileException:
            print("Error creating or reading master workbook")
            return False

    def proccessFile(self, file):
        #File Consolidation
        master = self.__getWorkbook(MASTER_PATH)    #get master workbook
        newFile = self.__getWorkbook(file)          #get file found workbook
        if (not master or not newFile):
            return
        try:
            sheet_number = len(master.worksheets)+1
            for sheet in newFile:
                name = "Sheet" + str(sheet_number)  #name of the new sheet
                master.create_sheet(name)
                master.save(MASTER_PATH)
                master.active = master[name]
                ws2 = master.active
                #iterate through the sheet of the new file
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        c = sheet.cell(row=i, column=j)
                        ws2.cell(row=i, column=j).value = c.value   #put value
                master.save(MASTER_PATH)
                sheet_number += 1
            return True
        except utils.exceptions.SheetTitleException:
            print('Error with Sheet Names '+file)
        except:
            print("Error consolidating File at"+file)

        return False
